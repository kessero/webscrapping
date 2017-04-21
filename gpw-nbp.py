#!/usr/bin/python2
import re
import json
import bs4
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import time

dest_filename = ('/home/kesser/MEGA/Python/nbp/przyklad.xlsx')
wbw = Workbook()
gpw_sheet = wbw.create_sheet(title='gpw')
nbp_sheet = wbw.create_sheet(title='nbp')
wbw.save(dest_filename)
text_data_gpw = []
tabela_kursow_gpw = {}
text_data_nbp = []
tabela_kursow_nbp = {}
column_names = []
nazwy_kolumn = {}
etykiety = []
etykiety_list = {}


def gpw():
    if __name__ == '__main__':
        respond = requests.get("https://www.gpw.pl/analizy_i_statystyki")

        soup = bs4.BeautifulSoup(respond.text, "html5lib")
        div = soup.find_all('div')
        divpa = soup.find(text=re.compile("stan na"))
    # wyciac ile sie da zbednych danych
    for table in soup.find_all('table'):
        for tr in table.find_all('tr'):
            tds = tr.find_all(['td'])
            for td in tds:
                text_data_gpw.append(td)

    # Handle column names if we find them
    css = soup.find_all("th", class_="w120px")
    for th in css:
        ths = th.get_text()
        column_names.append(ths)

        tabela_kursow_gpw['DATA'] = divpa  # dostep do slownika
        # cd. wycinania zakres danych, ktore mnie interesuja
    for i in range(0, 40):
        a = text_data_gpw[i].text.encode('ascii', 'ignore')
        a = (''.join(a.split()))
        b = a.replace(' ', '')
        b = b.strip()
        b = b.replace(',', '.')
        b = float(b)
        tabela_kursow_gpw[i] = b
    # otwarcie xls'a do edycji - trzeba drugi raz otworzyc ten plik do odczytu
    wba = load_workbook(dest_filename)
    # otwarcie dostepu do konkretnego skoroszytu
    wsa = wba['gpw']
    wsa.cell(column=1, row=1, value=divpa)  # wpisanie Stan na dzien
    wsa.cell(column=2, row=1, value=column_names[0]) # wpisanie nazwy "Krajowe"
    wsa.cell(column=3, row=1, value=column_names[1]) # wpisanie nazwy "Zagraniczne"
    wsa.cell(column=4, row=1, value=column_names[2]) # wpisanie nazwy "Razem"

    next_row = 2
    # zapisanie danych do odpowiednich kolumn i wierszy zakres musi byc mniejszy niz poprzednio ze wzgledu na inkrementacje w petli
    licznik = 0
    for i in range(0, 6):
        wsa.cell(column=2, row=next_row, value=tabela_kursow_gpw[licznik])
        wsa.cell(column=3, row=next_row, value=tabela_kursow_gpw[licznik + 1])
        wsa.cell(column=4, row=next_row, value=tabela_kursow_gpw[licznik + 2])
        next_row += 1
    # dodatkowy licznik pozwala przeskakiwac co 3 elementy w tabeli_kursow inaczej dubluja sie dane
        if licznik < 40:
           licznik += 3
    #licznik += 3

    wba.save(dest_filename)
def nbp():
    if __name__ == '__main__':
    	respond = requests.get("http://www.nbp.pl/home.aspx?f=/kursy/kursya.html")

    soup = bs4.BeautifulSoup(respond.text, "html5lib")
    div = soup.find_all('div')
    divpa = soup.find(text=re.compile("z dnia"))
    #wyciac ile sie da zbednych danych
    for table in soup.find_all('table'):
    	for tr in table.find_all('tr'):
            	tds = tr.find_all(['td'])
    		for td in tds:
    			text_data_nbp.append(td)
    tabela_kursow_nbp['DATA'] = divpa
    # cd. wycinania zakres danych ktore nas interesuja to 62-167
    for i in range(62, 167):
    	a = text_data_nbp[i].text
    	a = (' '.join(a))
    	tabela_kursow_nbp[i] = a
    #otwarcie xls'a do edycji - trzeba drugi raz otworzyc ten plik do odczytu
    wba=load_workbook(dest_filename)
    #otwarcie dostepu do konkretnego skoroszytu
    wsa=wba['nbp']
    wsa.cell(column=2, row=1, value=divpa)
    next_row = 2
    #zapisanie danych do odpowiednich kolumn i wierszy zakres musi byc mniejszy niz poprzednio ze wzgledu na inkrementacje w petli
    licznik = 62
    for i in range(62, 97):
    		wsa.cell(column=1, row=next_row, value=tabela_kursow_nbp[licznik])
    		wsa.cell(column=2, row=next_row, value=tabela_kursow_nbp[licznik+1])
    		wsa.cell(column=3, row=next_row, value=tabela_kursow_nbp[licznik+2])
    		next_row += 1
    # dodatkowy licznik pozwala przeskakiwac co 3 elementy w tabeli_kursow inaczej dubluja sie dane
    		if licznik <164:
    			licznik +=3

    wba.save(dest_filename)
gpw()
nbp()
