#!/usr/bin/python2
import re
import json
import bs4
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import time
dest_filename = ('/home/kesser/Python/przyklad.xlsx')
wbw = Workbook()
wsw = wbw.create_sheet(title='przyklad')
wbw.save(dest_filename)
text_data = []
tabela_kursow = {}
if __name__ == '__main__':
	respond = requests.get("http://www.nbp.pl/home.aspx?f=/kursy/kursya.html")

soup = bs4.BeautifulSoup(respond.text, "html5lib")
div = soup.find_all('div')
divpa = soup.find(text=re.compile("z dnia"))
#wyciac ile sie da zbednych danych
for table in soup.find_all('table'):
	for tr in table.find_all('tr'):			
        	tds = tr.find_all(['td'])		
		for td in tds:
			text_data.append(td)
tabela_kursow['DATA'] = divpa
# cd. wycinania zakres danych ktore nas interesuja to 62-167
for i in range(62, 167):
	a = text_data[i].text
	a = (' '.join(a))
	tabela_kursow[i] = a
#otwarcie xls'a do edycji - trzeba drugi raz otworzyc ten plik do odczytu
wba=load_workbook(dest_filename)
#otwarcie dostepu do konkretnego skoroszytu
wsa=wba['przyklad']
wsa.cell(column=2, row=1, value=divpa)
next_row = 2
#zapisanie danych do odpowiednich kolumn i wierszy zakres musi byc mniejszy niz poprzednio ze wzgledu na inkrementacje w petli
licznik = 62
for i in range(62, 97):
		wsa.cell(column=1, row=next_row, value=tabela_kursow[licznik])
		wsa.cell(column=2, row=next_row, value=tabela_kursow[licznik+1])
		wsa.cell(column=3, row=next_row, value=tabela_kursow[licznik+2])
		next_row += 1
# dodatkowy licznik pozwala przeskakiwac co 3 elementy w tabeli_kursow inaczej dubluja sie dane
		if licznik <164:		
			licznik +=3			
		
wba.save(dest_filename)
