#!/usr/bin/python2
import re
import json
import bs4
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import time

dest_filename = ('/home/kesser/MEGA/Python/nbp/przyklad.xlsx')
wbw = Workbook()
wsw = wbw.create_sheet(title='gpw')
wbw.save(dest_filename)
text_data = []
tabela_kursow = {}
column_names = []
nazwy_kolumn = {}
etykiety = []
etykiety_list = {}

if __name__ == '__main__':
    respond = requests.get("https://www.gpw.pl/analizy_i_statystyki")

soup = bs4.BeautifulSoup(respond.text, "html5lib")
div = soup.find_all('div')
divpa = soup.find(text=re.compile("stan na"))
# wyciac ile sie da zbednych danych
for table in soup.find_all('table'):
    for tr in table.find_all('tr'):
        tds = tr.find_all(['td'])
        for td in tds:
            text_data.append(td)
    
#    for tr in table.find_all('tr'):
#        thsa = tr.find_all(['th'])
#        for thssa in thsa:
#            etykiety.append(thsaa)

# Handle column names if we find them
css = soup.find_all("th", class_="w120px")
for th in css:
    ths = th.get_text() 
    column_names.append(ths)
		
tabela_kursow['DATA'] = divpa  # dostep do slownika 
# cd. wycinania zakres danych, ktore mnie interesuja 
for i in range(0, 40):
    a = text_data[i].text.encode('ascii', 'ignore')
    a = (''.join(a.split()))
    b = a.replace(' ', '')
    b = b.strip()
    b = b.replace(',', '.')
    b = float(b)
    tabela_kursow[i] = b
# otwarcie xls'a do edycji - trzeba drugi raz otworzyc ten plik do odczytu
wba = load_workbook(dest_filename)
# otwarcie dostepu do konkretnego skoroszytu
wsa = wba['gpw']
wsa.cell(column=1, row=1, value=divpa)  # wpisanie Stan na dzien
wsa.cell(column=2, row=1, value=column_names[0]) # wpisanie nazwy "Krajowe"
wsa.cell(column=3, row=1, value=column_names[1]) # wpisanie nazwy "Zagraniczne"
wsa.cell(column=4, row=1, value=column_names[2]) # wpisanie nazwy "Razem"

#wsa.cell(column=1, row=2, value=etykiety_list[0]) # wpisanie nazwy "Rynek podstawowy"
#wsa.cell(column=1, row=3, value=etykiety_list[1]) # wpisanie nazwy "Rynek rownolegly"
#wsa.cell(column=1, row=4, value=etykiety_list[2]) # wpisanie nazwy "Razem"

next_row = 2
# zapisanie danych do odpowiednich kolumn i wierszy zakres musi byc mniejszy niz poprzednio ze wzgledu na inkrementacje w petli
licznik = 0
for i in range(0, 6):
    wsa.cell(column=2, row=next_row, value=tabela_kursow[licznik])
    wsa.cell(column=3, row=next_row, value=tabela_kursow[licznik + 1])
    wsa.cell(column=4, row=next_row, value=tabela_kursow[licznik + 2])
    next_row += 1
# dodatkowy licznik pozwala przeskakiwac co 3 elementy w tabeli_kursow inaczej dubluja sie dane
    if licznik < 40:
       licznik += 3
#licznik += 3

wba.save(dest_filename)

